import datetime
import hashlib
import hmac
import json
import random
import secrets
from io import BytesIO
from time import sleep

import pandas as pd
import requests
from decouple import config
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from requests.adapters import HTTPAdapter
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from urllib3 import Retry
from django.conf import settings

from . import models, serializers
from .other_tests import tranx_id_generator
# Create your views here.

import firebase_admin
from firebase_admin import credentials
from firebase_admin import db, firestore
from secrets import compare_digest

if not firebase_admin._apps:
    cred = credentials.Certificate(settings.FIREBASE_ADMIN_CERT)
    firebase_admin.initialize_app(cred, {
        'databaseURL': 'https://bestpay-flutter-default-rtdb.firebaseio.com'
    })

database = firestore.client()
user_collection = database.collection(u'Users')
history_collection = database.collection(u'History Web')
mail_collection = database.collection('mail')
mtn_history = database.collection('MTN_Admin_History')
mtn_tranx = mtn_history.document('mtnTransactions')
big_time = mtn_tranx.collection('big_time')
mtn_other = mtn_tranx.collection('mtnOther')
bearer_token_collection = database.collection("_KeysAndBearer")
history_web = database.collection(u'History Web').document('all_users')


# all_users = [{**user.to_dict(), "id": user.id} for user in user_collection]
# print(all_users)
# user1 = user_collection.document('B2Ruk9G9b0ZX8mvluBUL3jKOCuA3')


# print(user1.get())
# print({**user1.to_dict(), "id": user1.id})


def all_users():
    """Get all todo from firestore database"""
    docs = user_collection.stream()
    for doc in docs:
        print(doc.to_dict())


def get_user_details(user_id):
    user = user_collection.document(user_id)
    doc = user.get()
    if doc.exists:
        doc_dict = doc.to_dict()
        first_name = doc_dict['first name']
        last_name = doc_dict['last name']
        email = doc_dict['email']
        phone = doc_dict['phone']
        print(first_name), print(last_name), print(email), print(phone)
        return doc.to_dict()
    else:
        return None


def update_user_wallet(user_id, amount):
    print("did updating")
    print(f"amount:{amount}")
    user = get_user_details(user_id)
    if user is None:
        return None
    user_wallet = user['wallet']
    new_balance = float(user_wallet) - float(amount)
    print(f"new_balance:{new_balance}")
    print(user_wallet)
    doc_ref = user_collection.document(user_id)
    doc_ref.update({'wallet': new_balance})
    user = get_user_details(user_id)
    user_wallet = user['wallet']
    print(f"new_user_wallet: {user_wallet}")
    if user_wallet is not None:
        return user_wallet
    else:
        return None


def tranx_id_gen():
    tranx_id = tranx_id_generator()
    history = history_collection.document(str(tranx_id))
    doc = history.get()
    if doc.exists:
        return tranx_id_gen()
    else:
        return str(tranx_id)


def check_user_balance_against_price(user_id, price):
    details = get_user_details(user_id)
    wallet_balance = details['wallet']
    if wallet_balance is not None:
        return wallet_balance >= float(price)
    else:
        return None


# History
def get_all_history():
    docs = history_collection.stream()
    counter = 0
    for doc in docs:
        if counter < 10:
            print(doc.to_dict())
            counter = counter + 1
            print(f"counter {counter}")
        break


def send_ishare_bundle(first_name: str, last_name: str, buyer, receiver: str, email: str, bundle: float,
                       reference: str):
    url = "https://console.bestpaygh.com/api/flexi/v1/new_transaction/"

    token = bearer_token_collection.document("Active_API_BoldAssure")
    token_doc = token.get()
    token_doc_dict = token_doc.to_dict()
    key = token_doc_dict['key']
    secret = token_doc_dict['secret']
    print(key)
    print(secret)

    headers = {
        "api-key": key,
        "api-secret": secret,
        'Content-Type': 'application/json'
    }

    payload = json.dumps({
        "first_name": first_name if first_name != "" else "First Name",
        "last_name": last_name if last_name != "" else "Last Name",
        "account_number": buyer,
        "receiver": receiver,
        "account_email": email,
        "reference": reference,
        "bundle_amount": bundle
    })

    print(payload)

    session = requests.Session()
    retry = Retry(connect=15, backoff_factor=0.5)
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('https://', adapter)

    response = session.post(url, headers=headers, data=payload)
    status_code = response.status_code
    print("after response")

    return response, status_code


def send_and_save_to_history(user_id, txn_type: str, txn_status: str, paid_at: str, ishare_balance: float,
                             color_code: str,
                             data_volume: float, reference: str, data_break_down: str, amount: float, receiver: str,
                             date: str, image, time: str, date_and_time: str):
    user_details = get_user_details(user_id)
    first_name = user_details['first name']
    last_name = user_details['last name']
    email = user_details['email']
    phone = user_details['phone']
    wallet = user_details['wallet']

    data = {
        'batch_id': "unknown",
        'buyer': phone,
        'color_code': color_code,
        'amount': amount,
        'data_break_down': data_break_down,
        'data_volume': data_volume,
        'date': date,
        'date_and_time': date_and_time,
        'done': "unknown",
        'email': email,
        'image': user_id,
        'ishareBalance': ishare_balance,
        'name': f"{first_name} {last_name}",
        'number': receiver,
        'paid_at': paid_at,
        'reference': reference,
        'responseCode': "0",
        'status': txn_status,
        'time': time,
        'tranxId': str(tranx_id_gen()),
        'type': txn_type,
        'uid': user_id,
        'bal': wallet
    }
    history_collection.document(date_and_time).set(data)
    history_web.collection(email).document(date_and_time).set(data)

    print("first save")

    ishare_response, status_code = send_ishare_bundle(first_name=first_name, last_name=last_name, receiver=receiver,
                                                      buyer=phone,
                                                      bundle=data_volume,
                                                      email=email, reference=reference)
    json_response = ishare_response.json()
    print(f"hello:{json_response}")
    status_code = status_code
    print(status_code)
    batch_id = json_response["batch_id"]
    print(batch_id)

    doc_ref = history_collection.document(date_and_time)
    doc_ref.update({'batch_id': batch_id, 'responseCode': status_code})
    history_web.collection(email).document(date_and_time).update({'batch_id': batch_id, 'responseCode': status_code})
    # data = {
    #     'batch_id': batch_id,
    #     'buyer': phone,
    #     'color_code': color_code,
    #     'amount': amount,
    #     'data_break_down': data_break_down,
    #     'data_volume': data_volume,
    #     'date': date,
    #     'date_and_time': date_and_time,
    #     'done': "unknown",
    #     'email': email,
    #     'image': image,
    #     'ishareBalance': ishare_balance,
    #     'name': f"{first_name} {last_name}",
    #     'number': receiver,
    #     'paid_at': paid_at,
    #     'reference': reference,
    #     'responseCode': status_code,
    #     'status': txn_status,
    #     'time': time,
    #     'tranxId': str(tranx_id_gen()),
    #     'type': txn_type,
    #     'uid': user_id
    # }
    # history_collection.document(date_and_time).set(data)
    # history_web.collection(email).document(date_and_time).set(data)

    print("firebase saved")
    return status_code, batch_id if batch_id else "No batchId", email, first_name


def ishare_verification(batch_id):
    if batch_id == "No batchId":
        return False

    url = f"https://backend.boldassure.net:445/live/api/context/business/airteltigo-gh/ishare/tranx-status/{batch_id}"

    payload = {}
    token = bearer_token_collection.document("Active_API_BoldAssure")
    token_doc = token.get()
    token_doc_dict = token_doc.to_dict()
    tokennn = token_doc_dict['ishare_bearer']
    headers = {
        'Authorization': tokennn
    }

    response = requests.request("GET", url, headers=headers, data=payload)

    if response.status_code == 200:
        json_data = response.json()
        print(json_data)
        return json_data
    else:
        return False


# all_users()

# print(get_user_wallet('9VA0qyq6lXYPZ6Ut867TVcBvF2t1'))
# print(get_user_details('9VA0qyq6lXYPZ6Ut867TVcBvF2t1'))
# print(check_user_balance_against_price('9VA0qyq6lXYPZ6Ut867TVcBvF2t1', 857585758))


# get_all_history()


class WalletUserListApiView(APIView):

    def get(self, *args, **kwargs):
        # wallet_users = users_ref.get()
        # serializer = serializers.WalletUserSerializer(wallet_users, many=True)
        # return Response(wallet_users, status=status.HTTP_200_OK)
        ...

    def post(self, request, *args, **kwargs):
        data = {
            'user_id': request.data.get('user_id'),
            'wallet_balance': request.data.get('wallet_balance'),
            'status': request.data.get('status')
        }
        print(data)
        serializer = serializers.WalletUserSerializer(data=data)
        try:
            converted = float(data['wallet_balance'])
        except ValueError:
            return Response({'message': "Invalid wallet balance provided"}, status=status.HTTP_400_BAD_REQUEST)
        user_status = data['status']
        if user_status not in ["Active", "Inactive"]:
            return Response({'message': "Invalid status provided. Valid options are 'Active' or 'Inactive'."},
                            status=status.HTTP_400_BAD_REQUEST)
        # if users_ref.child(data['user_id']):
        #     return Response({"code": "0001", "message": "User wallet already exists"}, status=status.HTTP_400_BAD_REQUEST)
        #
        # if serializer.is_valid():
        #     users_ref.child(data['user_id']).set(
        #         {
        #             'user_id': data['user_id'],
        #             'wallet_balance': data['wallet_balance'],
        #             'status': data['status']
        #         }
        #     )
        #     return Response({"code": "0000", "message": "User wallet created successfully", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class WalletUserDetail(APIView):

    def get_object(self, user_id):
        try:
            # user = users_ref.child(user_id)
            user = "hi"
            return user
        except:
            return None

    def get(self, request, user_id, *args, **kwargs):
        user_instance = self.get_object(user_id)
        if not user_instance:
            return Response({
                "message": "User does not exist"
            }, status=status.HTTP_404_NOT_FOUND)
        # serializer = user_instance.get()
        # user_dict = dict(serializer)
        # wallet = user_dict["wallet_balance"]
        # print(wallet)
        # print(user_dict)
        # print(serializer)
        # return Response({'user_details': serializer}, status=status.HTTP_200_OK)
        ...

    def post(self, request, user_id, *args, **kwargs):
        user_instance = self.get_object(user_id)
        if not user_instance:
            return Response({
                "message": "User does not exist"
            }, status=status.HTTP_404_NOT_FOUND)
        data = {
            'user_id': request.data.get('user_id'),
            'wallet_balance': request.data.get('wallet_balance'),
            'status': request.data.get('status')
        }
        print(data)
        user_data = user_instance.get()
        serializer = serializers.WalletUserSerializer(data=data)
        try:
            converted = float(data['wallet_balance'])
        except ValueError:
            return Response({'message': "Invalid wallet balance provided"}, status=status.HTTP_400_BAD_REQUEST)
        user_status = data['status']
        if user_status not in ["Active", "Inactive"]:
            return Response({'message': "Invalid status provided. Valid options are 'Active' or 'Inactive'."},
                            status=status.HTTP_400_BAD_REQUEST)

        # if serializer.is_valid():
        #     users_ref.child(user_id).update(
        #         {
        #             'wallet_balance': data['wallet_balance'],
        #             'user_id': data['user_id'],
        #             'status': data['status'],
        #         }
        #     )
        #     new_user_details = users_ref.child(user_id).get()
        #     return Response({"code": "0000", "message": "User details updated successfully", "data": new_user_details},
        #                     status=status.HTTP_200_OK)
        # return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class WalletUserBalance(APIView):
    def get_object(self, user_id):
        try:
            # user = users_ref.child(user_id)
            user = "hi"
            return user
        except TypeError:
            return None

    # def get(self, request, user_id, *args, **kwargs):
    #     user_instance = self.get_object(user_id)
    #     if not user_instance:
    #         return Response({
    #             "message": "User does not exist"
    #         }, status=status.HTTP_404_NOT_FOUND)
    # serializer = user_instance.get()
    # try:
    #     user_dict = dict(serializer)
    # except TypeError:
    #     return Response({'code': '0001', 'message': "User not found"}, status=status.HTTP_400_BAD_REQUEST)
    # wallet = user_dict["wallet_balance"]
    # return Response({'code': '0000', 'wallet_balance': wallet}, status=status.HTTP_200_OK)

    def get(self, request, token, user_id, amount: str, txn_type: str, txn_status: str, paid_at: str,
            channel: str, ishare_balance: str,
            color_code: str,
            data_volume: str, reference: str, data_break_down: str, receiver: str,
            date: str, image, time: str, date_and_time: str, callback_url, *args, **kwargs):
        if token != config('TOKEN'):
            return Response(data={'message': 'Invalid Authorization Token Provided'},
                            status=status.HTTP_401_UNAUTHORIZED)
        print("hit it")
        user_details = get_user_details(user_id)
        email = user_details['email']
        user_instance = self.get_object(user_id)
        # hist = history_web.collection(email).document(date_and_time)
        # doc = hist.get()
        # if doc.exists:
        #     print(doc)
        #     return redirect(f"https://{callback_url}")
        # else:
        #     print("no record found")
        if not user_instance:
            return Response({
                "message": "User does not exist"
            }, status=status.HTTP_404_NOT_FOUND)
        data = {
            'user_id': user_id,
            'top_up_amount': amount,
        }
        user_details = get_user_details(user_id)
        first_name = user_details['first name']
        last_name = user_details['last name']
        phone = user_details['phone']
        to_be_added = float(amount)
        all_data = {
            'batch_id': "unknown",
            'buyer': phone,
            'color_code': color_code,
            'amount': amount,
            'data_break_down': data_break_down,
            'data_volume': data_volume,
            'date': date,
            'date_and_time': date_and_time,
            'done': "Success",
            'email': email,
            'image': user_id,
            'ishareBalance': ishare_balance,
            'name': f"{first_name} {last_name}",
            'number': receiver,
            'paid_at': paid_at,
            'reference': reference,
            'responseCode': 200,
            'status': txn_status,
            'time': time,
            'tranxId': str(tranx_id_gen()),
            'type': txn_type,
            'uid': user_id
        }
        history_web.collection(email).document(date_and_time).set(all_data)
        print("saved")
        history_collection.document(date_and_time).set(all_data)
        print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
        print("saved")
        print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")
        print(data)
        user_details = get_user_details(user_id)
        try:
            previous_wallet_balance = user_details['wallet']
        except ValueError:
            return Response({'code': '0001', 'message': "User not found"}, status=status.HTTP_400_BAD_REQUEST)
        serializer = serializers.WalletUserSerializer(data=data)
        if serializer.is_valid():
            user_details = get_user_details(user_id)
            first_name = user_details['first name']
            last_name = user_details['last name']
            email = user_details['email']
            phone = user_details['phone']
            to_be_added = float(amount)
            print(to_be_added)
            print("heyyyyyyyy")
            new_balance = previous_wallet_balance + to_be_added
            print(new_balance)
            doc_ref = user_collection.document(user_id)
            doc_ref.update(
                {'wallet': new_balance, 'wallet_last_update': date_and_time, 'recent_wallet_reference': reference})
            print(doc_ref.get().to_dict())

            all_data = {
                'batch_id': "unknown",
                'buyer': phone,
                'color_code': color_code,
                'amount': amount,
                'data_break_down': data_break_down,
                'data_volume': data_volume,
                'date': date,
                'date_and_time': date_and_time,
                'done': "Success",
                'email': email,
                'image': user_id,
                'ishareBalance': ishare_balance,
                'name': f"{first_name} {last_name}",
                'number': receiver,
                'paid_at': paid_at,
                'reference': reference,
                'responseCode': 200,
                'status': txn_status,
                'time': time,
                'tranxId': str(tranx_id_gen()),
                'type': txn_type,
                'uid': user_id
            }
            history_web.collection(email).document(date_and_time).set(all_data)
            print("saved")
            history_collection.document(date_and_time).set(all_data)
            print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
            print("saved")
            print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")

            name = f"{first_name} {last_name}"
            amount = to_be_added
            file_path = 'wallet_api_app/wallet_mail.txt'
            mail_doc_ref = mail_collection.document()

            with open(file_path, 'r') as file:
                html_content = file.read()

            placeholders = {
                '{name}': name,
                '{amount}': amount
            }

            for placeholder, value in placeholders.items():
                html_content = html_content.replace(placeholder, str(value))

            mail_doc_ref.set({
                'to': email,
                'message': {
                    'subject': 'Wallet Topup',
                    'html': html_content,
                    'messageId': 'Bestpay'
                }
            })

            sms_message = f"GHS {to_be_added} was deposited in your wallet. Available balance is now GHS {new_balance}"
            sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to=0{user_details['phone']}&from=Bestpay&sms={sms_message}"
            response = requests.request("GET", url=sms_url)
            print(response.status_code)
            return redirect(f"https://{callback_url}")
        return redirect(f"https://{callback_url}")


class InitiateTransaction(APIView):
    def get(self, request, token, user_id: str, txn_type: str, txn_status: str, paid_at: str,
            channel: str, ishare_balance: str,
            color_code: str,
            data_volume: str, reference: str, data_break_down: str, amount: str, receiver: str,
            date: str, image, time: str, date_and_time: str, callback_url: str):
        if token != config('TOKEN'):
            return Response(data={'message': 'Invalid Authorization Token Provided'},
                            status=status.HTTP_401_UNAUTHORIZED)
        if channel.lower() == "wallet":
            enough_balance = check_user_balance_against_price(user_id, amount)
        else:
            enough_balance = True
            print("not wallet")
        print(enough_balance)
        if enough_balance:
            user_details = get_user_details(user_id)
            email = user_details['email']
            print(enough_balance)
            # hist = history_web.collection(email).document(date_and_time)
            # doc = hist.get()
            # if doc.exists:
            #     return redirect(f"https://{callback_url}")
            # else:
            #     print("no record found")
            if channel.lower() == "wallet":
                user = get_user_details(user_id)
                if user is None:
                    return None
                previous_user_wallet = user['wallet']
                print(f"previous wallet: {previous_user_wallet}")
                new_balance = float(previous_user_wallet) - float(amount)
                print(f"new_balance:{new_balance}")
                doc_ref = user_collection.document(user_id)
                doc_ref.update({'wallet': new_balance})
                user = get_user_details(user_id)
                new_user_wallet = user['wallet']
                print(f"new_user_wallet: {new_user_wallet}")
                if new_user_wallet == previous_user_wallet:
                    user = get_user_details(user_id)
                    if user is None:
                        return None
                    previous_user_wallet = user['wallet']
                    print(f"previous wallet: {previous_user_wallet}")
                    new_balance = float(previous_user_wallet) - float(amount)
                    print(f"new_balance:{new_balance}")
                    doc_ref = user_collection.document(user_id)
                    doc_ref.update({'wallet': new_balance})
                    user = get_user_details(user_id)
                    new_user_wallet = user['wallet']
                    print(f"new_user_wallet: {new_user_wallet}")
                else:
                    print("it's fine")
            status_code, batch_id, email, first_name = send_and_save_to_history(user_id, txn_type, txn_status, paid_at,
                                                                                float(ishare_balance),
                                                                                color_code, float(data_volume),
                                                                                reference,
                                                                                data_break_down,
                                                                                float(amount), receiver,
                                                                                date, image, time, date_and_time)
            print(status_code)
            print(batch_id)
            sleep(10)
            ishare_verification_response = ishare_verification(batch_id)
            if ishare_verification_response is not False:
                code = \
                    ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                        "apiResponse"][
                        "responseCode"]
                ishare_response = \
                    ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                        "ishareApiResponseData"][
                        "apiResponseData"][
                        0][
                        "responseMsg"]
                print(code)
                print(ishare_response)
                if code == '200' or ishare_response == 'Crediting Successful.':
                    sms = f"Hey there\nYour account has been credited with {data_volume}MB.\nConfirm your new balance using the AT Mobile App"
                    r_sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={receiver}&from=Bestpay&sms={sms}"
                    response = requests.request("GET", url=r_sms_url)
                    print(response.text)
                    doc_ref = history_collection.document(date_and_time)
                    doc_ref.update({'done': 'Successful'})
                    mail_doc_ref = mail_collection.document(f"{batch_id}-Mail")
                    file_path = 'wallet_api_app/mail.txt'  # Replace with your file path

                    name = first_name
                    volume = data_volume
                    date = date_and_time
                    reference_t = reference
                    receiver_t = receiver

                    with open(file_path, 'r') as file:
                        html_content = file.read()

                    placeholders = {
                        '{name}': name,
                        '{volume}': volume,
                        '{date}': date,
                        '{reference}': reference_t,
                        '{receiver}': receiver_t
                    }

                    for placeholder, value in placeholders.items():
                        html_content = html_content.replace(placeholder, str(value))

                    mail_doc_ref.set({
                        'to': email,
                        'message': {
                            'subject': 'AT Flexi Bundle',
                            'html': html_content,
                            'messageId': 'Bestpay'
                        }
                    })
                else:
                    doc_ref = history_collection.document(date_and_time)
                    doc_ref.update({'done': 'Failed'})
                user = history_collection.document(date_and_time)
                doc = user.get()
                print(doc.to_dict())
                # return Response(data={'status_code': status_code, 'batch_id': batch_id}, status=status.HTTP_200_OK)
                return redirect(f"https://{callback_url}")
            else:
                return redirect(f"https://{callback_url}")
        else:
            return redirect(f"https://{callback_url}")
            # return Response({"code": '0001', 'message': 'Not enough balance to perform transaction'},
            #                 status=status.HTTP_200_OK)


# class InitiateMTNTransaction(APIView):
#     ...


class InitiateBigTimeTransaction(APIView):
    def get(self, request, token, user_id: str, txn_type: str, txn_status: str, paid_at: str,
            channel: str, ishare_balance: str,
            color_code: str,
            data_volume: str, reference: str, data_break_down: str, amount: str, receiver: str,
            date: str, image, time: str, date_and_time: str, callback_url: str):
        if token != config('TOKEN'):
            return Response(data={'message': 'Invalid Authorization Token Provided'},
                            status=status.HTTP_401_UNAUTHORIZED)
        if channel.lower() == "wallet":
            print("used this")
            enough_balance = check_user_balance_against_price(user_id, amount)
        else:
            enough_balance = True
        print(enough_balance)
        if enough_balance:
            user_details = get_user_details(user_id)
            first_name = user_details['first name']
            last_name = user_details['last name']
            email = user_details['email']
            phone = user_details['phone']
            # hist = history_web.collection(email).document(date_and_time)
            # doc = hist.get()
            # if doc.exists:
            #     print(doc)
            #     return redirect(f"https://{callback_url}")
            # else:
            #     print("no record found")
            if channel.lower() == "wallet":
                print("updated")
                user = get_user_details(user_id)
                if user is None:
                    return None
                previous_user_wallet = user['wallet']
                print(f"previous wallet: {previous_user_wallet}")
                new_balance = float(previous_user_wallet) - float(amount)
                print(f"new_balance:{new_balance}")
                doc_ref = user_collection.document(user_id)
                doc_ref.update({'wallet': new_balance})
                user = get_user_details(user_id)
                new_user_wallet = user['wallet']
                print(f"new_user_wallet: {new_user_wallet}")
                if new_user_wallet == previous_user_wallet:
                    user = get_user_details(user_id)
                    if user is None:
                        return None
                    previous_user_wallet = user['wallet']
                    print(f"previous wallet: {previous_user_wallet}")
                    new_balance = float(previous_user_wallet) - float(amount)
                    print(f"new_balance:{new_balance}")
                    doc_ref = user_collection.document(user_id)
                    doc_ref.update({'wallet': new_balance})
                    user = get_user_details(user_id)
                    new_user_wallet = user['wallet']
                    print(f"new_user_wallet: {new_user_wallet}")
                else:
                    print("it's fine")
            user = get_user_details(user_id)
            bal = user['wallet']
            data = {
                'batch_id': "unknown",
                'buyer': phone,
                'color_code': color_code,
                'amount': amount,
                'data_break_down': data_break_down,
                'data_volume': data_volume,
                'date': date,
                'date_and_time': date_and_time,
                'done': "unknown",
                'email': email,
                'image': user_id,
                'ishareBalance': ishare_balance,
                'name': f"{first_name} {last_name}",
                'number': receiver,
                'paid_at': paid_at,
                'reference': reference,
                'responseCode': 200,
                'status': txn_status,
                'time': time,
                'bal': bal,
                'tranxId': str(tranx_id_gen()),
                'type': txn_type,
                'uid': user_id
            }
            history_collection.document(date_and_time).set(data)
            history_web.collection(email).document(date_and_time).set(data)
            big_time.document(date_and_time).set(data)
            user = history_collection.document(date_and_time)
            doc = user.get()
            print(doc.to_dict())
            tranx_id = doc.to_dict()['tranxId']
            mail_doc_ref = mail_collection.document()
            file_path = 'wallet_api_app/mtn_mail.txt'  # Replace with your file path

            name = first_name
            volume = data_volume
            date = date_and_time
            reference_t = reference
            receiver_t = receiver

            with open(file_path, 'r') as file:
                html_content = file.read()

            placeholders = {
                '{name}': name,
                '{volume}': volume,
                '{date}': date,
                '{reference}': reference_t,
                '{receiver}': receiver_t
            }

            for placeholder, value in placeholders.items():
                html_content = html_content.replace(placeholder, str(value))

            mail_doc_ref.set({
                'to': email,
                'message': {
                    'subject': 'Big Time Data',
                    'html': html_content,
                    'messageId': 'Bestpay'
                }
            })
            return redirect(f"https://{callback_url}")
        else:
            return Response({"code": '0001', 'message': 'Not enough balance to perform transaction'},
                            status=status.HTTP_200_OK)


class InitiateMTNTransaction(APIView):
    def get(self, request, token, user_id: str, txn_type: str, txn_status: str, paid_at: str,
            channel: str, ishare_balance: str,
            color_code: str,
            data_volume: str, reference: str, data_break_down: str, amount: str, receiver: str,
            date: str, image, time: str, date_and_time: str, callback_url: str):
        if token != config('TOKEN'):
            return Response(data={'message': 'Invalid Authorization Token Provided'},
                            status=status.HTTP_401_UNAUTHORIZED)
        if channel.lower() == "wallet":
            print("used this")
            enough_balance = check_user_balance_against_price(user_id, amount)
        else:
            enough_balance = True
            print("not wallet")
        print(enough_balance)
        if enough_balance:
            user_details = get_user_details(user_id)
            first_name = user_details['first name']
            last_name = user_details['last name']
            email = user_details['email']
            phone = user_details['phone']
            # hist = history_web.collection(email).document(date_and_time)
            # doc = hist.get()
            # if doc.exists:
            #     print(doc)
            #     return redirect(f"https://{callback_url}")
            # else:
            #     print("no record found")
            if channel.lower() == "wallet":
                print("updated")
                user = get_user_details(user_id)
                if user is None:
                    return None
                previous_user_wallet = user['wallet']
                print(f"previous wallet: {previous_user_wallet}")
                new_balance = float(previous_user_wallet) - float(amount)
                print(f"new_balance:{new_balance}")
                doc_ref = user_collection.document(user_id)
                doc_ref.update({'wallet': new_balance})
                user = get_user_details(user_id)
                new_user_wallet = user['wallet']
                print(f"new_user_wallet: {new_user_wallet}")
                if new_user_wallet == previous_user_wallet:
                    user = get_user_details(user_id)
                    if user is None:
                        return None
                    previous_user_wallet = user['wallet']
                    print(f"previous wallet: {previous_user_wallet}")
                    new_balance = float(previous_user_wallet) - float(amount)
                    print(f"new_balance:{new_balance}")
                    doc_ref = user_collection.document(user_id)
                    doc_ref.update({'wallet': new_balance})
                    user = get_user_details(user_id)
                    new_user_wallet = user['wallet']
                    print(f"new_user_wallet: {new_user_wallet}")
                else:
                    print("it's fine")
            user = get_user_details(user_id)
            bal = user['wallet']
            data = {
                'batch_id': "unknown",
                'buyer': phone,
                'color_code': color_code,
                'amount': amount,
                'data_break_down': data_break_down,
                'data_volume': data_volume,
                'date': date,
                'date_and_time': date_and_time,
                'done': "unknown",
                'email': email,
                'image': user_id,
                'ishareBalance': ishare_balance,
                'name': f"{first_name} {last_name}",
                'number': receiver,
                'paid_at': paid_at,
                'reference': reference,
                'responseCode': 200,
                'status': txn_status,
                'time': time,
                'bal': bal,
                'tranxId': str(tranx_id_gen()),
                'type': txn_type,
                'uid': user_id
            }
            history_collection.document(date_and_time).set(data)
            history_web.collection(email).document(date_and_time).set(data)
            user = history_collection.document(date_and_time)
            doc = user.get()
            hist = history_collection.document(date_and_time)
            new_mtn_txn = models.MTNTransaction.objects.create(
                user_id=user_id,
                amount=amount,
                bundle_volume=data_volume,
                number=receiver,
                firebase_date=date_and_time
            )
            new_mtn_txn.save()
            print("first")
            print(hist.get().to_dict())
            print(doc.to_dict())
            tranx_id = doc.to_dict()['tranxId']
            user = get_user_details(user_id)
            bal = user['wallet']
            second_data = {
                'amount': amount,
                'batch_id': "unknown",
                'channel': channel,
                'color_code': color_code,
                'created_at': date_and_time,
                'data_volume': data_volume,
                'date': date,
                'email': email,
                'date_and_time': date_and_time,
                'image': user_id,
                'ip_address': "",
                'ishareBalance': 0,
                'name': f"{first_name} {last_name}",
                'number': receiver,
                'buyer': phone,
                'paid_at': date_and_time,
                'payment_status': "success",
                'reference': reference,
                'status': "Undelivered",
                'time': time,
                'bal': bal,
                'tranxId': tranx_id,
                'type': "MTN Master Bundle"
            }
            mtn_other.document(date_and_time).set(second_data)
            user22 = mtn_other.document(date_and_time)
            pu = user22.get()
            print(pu.to_dict())
            print("pu")
            mail_doc_ref = mail_collection.document()
            file_path = 'wallet_api_app/mtn_maill.txt'  # Replace with your file path

            name = first_name
            volume = data_volume
            date = date_and_time
            reference_t = reference
            receiver_t = receiver

            with open(file_path, 'r') as file:
                html_content = file.read()

            placeholders = {
                '{name}': name,
                '{volume}': volume,
                '{date}': date,
                '{reference}': reference_t,
                '{receiver}': receiver_t
            }

            for placeholder, value in placeholders.items():
                html_content = html_content.replace(placeholder, str(value))

            mail_doc_ref.set({
                'to': email,
                'message': {
                    'subject': 'MTN Data',
                    'html': html_content,
                    'messageId': 'Bestpay'
                }
            })
            print("got to redirect")
            return redirect(f"https://{callback_url}")
        else:
            return Response({"code": '0001', 'message': 'Not enough balance to perform transaction'},
                            status=status.HTTP_200_OK)


# ============================================================================================================================

class MTNFlexiInitiate(APIView):
    def post(self, request):
        token = request.headers.get('Authorization')
        required_params = ['user_id', 'receiver', 'reference', 'data_volume', 'amount', 'channel']
        prices_dict = {
            1000: 3.9,
            2000: 7.8,
            3000: 11,
            4000: 14.5,
            5000: 18,
            6000: 21,
            7000: 24.5,
            8000: 27,
            10000: 32,
            15000: 48,
            20000: 64,
            25000: 78,
            30000: 94,
            40000: 128,
            50000: 155,
            100000: 290
        }
        # Check if the token matches the one in the environment variable
        if token != config("AT"):
            # Token matches, allow access
            return Response({'code': '0001', 'message': 'Unauthorized'}, status=status.HTTP_401_UNAUTHORIZED)
        else:
            request_data = request.data
            data = {
                "user_id": request.data.get('user_id'),
                "receiver": request.data.get('receiver'),
                "data_volume": request.data.get('data_volume'),
                "reference": request.data.get('reference'),
                "amount": request.data.get('amount'),
                "channel": request.data.get('channel'),
            }

            missing_params = [param for param in required_params if param not in request_data]

            if missing_params:
                print(missing_params)
                # If any required parameter is missing, return an error response
                return Response({'code': '0001', 'message': f'Missing parameters: {", ".join(missing_params)}'},
                                status=status.HTTP_400_BAD_REQUEST)
            else:
                user_id = data['user_id']
                receiver = data['receiver']
                data_volume = data['data_volume']
                reference = data['reference']
                amount = data['amount']
                amount_to_be_deducted = prices_dict[data_volume]
                channel = data['channel']
                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()
                if "wallet" == "wallet":
                    print("used this")
                    try:
                        enough_balance = check_user_balance_against_price(data['user_id'], amount_to_be_deducted)
                    except:
                        return Response({'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}'},
                                        status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(data['user_id'])
                    first_name = user_details['first name']
                    last_name = user_details['last name']
                    email = user_details['email']
                    phone = user_details['phone']
                    bal = user_details['wallet']
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     print(doc)
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        print("updated")
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount_to_be_deducted)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount_to_be_deducted)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")

                    data = {
                        'batch_id': "unknown",
                        'buyer': channel,
                        'color_code': "Green",
                        'amount': amount_to_be_deducted,
                        'data_break_down': data_volume,
                        'data_volume': data_volume,
                        'date': str(date),
                        'date_and_time': str(date_and_time),
                        'done': "unknown",
                        'email': email,
                        'image': user_id,
                        'ishareBalance': '',
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'paid_at': str(date_and_time),
                        'reference': reference,
                        'responseCode': 200,
                        'status': "Undelivered",
                        'bal': bal,
                        'time': str(time),
                        'tranxId': str(tranx_id_gen()),
                        'type': "MTN Master Bundle",
                        'uid': user_id
                    }

                    history_collection.document(date_and_time).set(data)
                    history_web.collection(email).document(date_and_time).set(data)
                    user = history_collection.document(date_and_time)
                    loko = history_web.collection(email).document(date_and_time)
                    new_mtn_txn = models.MTNTransaction.objects.create(
                        user_id=user_id,
                        amount=amount,
                        bundle_volume=data_volume,
                        number=receiver,
                        firebase_date=date_and_time
                    )
                    new_mtn_txn.save()
                    print(loko.get().to_dict())
                    doc = user.get()
                    print(doc.to_dict())
                    tranx_id = doc.to_dict()['tranxId']
                    second_data = {
                        'amount': amount_to_be_deducted,
                        'batch_id': "unknown",
                        'channel': "wallet",
                        'color_code': "Green",
                        'created_at': date_and_time,
                        'data_volume': data_volume,
                        'date': str(date),
                        'email': email,
                        'date_and_time': date_and_time,
                        'image': user_id,
                        'ip_address': "",
                        'ishareBalance': 0,
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'buyer': channel,
                        'paid_at': date_and_time,
                        'payment_status': "success",
                        'reference': reference,
                        'status': "Undelivered",
                        'bal': bal,
                        'time': str(time),
                        'tranxId': tranx_id,
                        'type': "MTN Master Bundle"
                    }
                    mtn_other.document(date_and_time).set(second_data)
                    user22 = mtn_other.document(date_and_time)
                    pu = user22.get()
                    print(pu.to_dict())
                    print("pu")
                    mail_doc_ref = mail_collection.document()
                    file_path = 'wallet_api_app/mtn_maill.txt'  # Replace with your file path

                    name = first_name
                    volume = data_volume
                    date = date_and_time
                    reference_t = reference
                    receiver_t = receiver

                    with open(file_path, 'r') as file:
                        html_content = file.read()

                    placeholders = {
                        '{name}': name,
                        '{volume}': volume,
                        '{date}': date,
                        '{reference}': reference_t,
                        '{receiver}': receiver_t
                    }

                    for placeholder, value in placeholders.items():
                        html_content = html_content.replace(placeholder, str(value))

                    mail_doc_ref.set({
                        'to': email,
                        'message': {
                            'subject': 'MTN Data',
                            'html': html_content,
                            'messageId': 'Bestpay'
                        }
                    })
                    print("got to redirect")
                    return Response(data={"code": "0000", "message": "Transaction saved"}, status=status.HTTP_200_OK)
                else:
                    return Response({"code": '0001', 'message': 'Not enough balance to perform transaction'},
                                    status=status.HTTP_400_BAD_REQUEST)


# ****************************************************************************************************
# ****************************************************************************************************

def webhook_send_and_save_to_history(user_id, txn_type: str, paid_at: str, ishare_balance: float,
                                     color_code: str,
                                     data_volume: float, reference: str, amount: float,
                                     receiver: str,
                                     date: str, time: str, date_and_time: str, txn_status):
    user_details = get_user_details(user_id)
    first_name = user_details['first name']
    last_name = user_details['last name']
    email = user_details['email']
    phone = user_details['phone']

    doc_ref = history_web.collection(email).document(date_and_time)
    if doc_ref.get().exists:
        data = doc_ref.get()
        data_dict = data.to_dict()
        batch_id = data_dict["batch_id"]
        if batch_id != "unknown":
            ishare_verification_response = ishare_verification(batch_id)
            if ishare_verification_response is not False:
                code = \
                    ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                        "apiResponse"][
                        "responseCode"]
                ishare_response = \
                    ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                        "ishareApiResponseData"][
                        "apiResponseData"][
                        0][
                        "responseMsg"]
                print(code)
                print(ishare_response)
                if code == '200' or ishare_response == 'Crediting Successful.':
                    return Response(data={"code": "0002"}, status=status.HTTP_200_OK)
                else:
                    pass
        else:
            pass
    else:
        pass
    print("moving on")
    data = {
        'batch_id': "unknown",
        'buyer': phone,
        'color_code': color_code,
        'amount': amount,
        'data_break_down': str(data_volume),
        'data_volume': data_volume,
        'date': date,
        'date_and_time': date_and_time,
        'done': "Pending",
        'email': email,
        'image': user_id,
        'ishareBalance': ishare_balance,
        'name': f"{first_name} {last_name}",
        'number': receiver,
        'paid_at': paid_at,
        'reference': reference,
        'responseCode': "0",
        'status': txn_status,
        'time': time,
        'tranxId': str(tranx_id_gen()),
        'type': txn_type,
        'uid': user_id
    }
    history_collection.document(date_and_time).set(data)
    history_web.collection(email).document(date_and_time).set(data)

    if history_collection.document(date_and_time).get().exists:
        print("first save")

    ishare_response, status_code = send_ishare_bundle(first_name=first_name, last_name=last_name, receiver=receiver,
                                                      buyer=phone,
                                                      bundle=data_volume,
                                                      email=email, reference=reference)
    json_response = ishare_response.json()
    print(f"hello:{json_response}")
    status_code = status_code
    print(status_code)
    try:
        batch_id = json_response["batch_id"]
    except KeyError:
        batch_id = "unknown"
    print(batch_id)

    doc_ref = history_collection.document(date_and_time)
    if doc_ref.get().exists:
        doc_ref.update({'batch_id': batch_id, 'responseCode': status_code})
        history_web.collection(email).document(date_and_time).update(
            {'batch_id': batch_id, 'responseCode': status_code})
    else:
        print("didn't find any entry to update")
    print("firebase saved")
    # return status_code, batch_id if batch_id else "No batchId", email, first_name
    return Response(
        data=json_response,
        status=status.HTTP_200_OK)


def hubtel_webhook_send_and_save_to_history(saved_data, user_id, reference, receiver, data_volume):
    user_details = get_user_details(user_id)
    first_name = user_details['first name']
    last_name = user_details['last name']
    email = user_details['email']
    phone = user_details['phone']

    doc_ref = history_web.collection(email).document(reference)
    if doc_ref.get().exists:
        data = doc_ref.get()
        data_dict = data.to_dict()
        batch_id = data_dict["batch_id"]
        if batch_id != "unknown":
            ishare_verification_response = ishare_verification(batch_id)
            if ishare_verification_response is not False:
                code = \
                    ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                        "apiResponse"][
                        "responseCode"]
                ishare_response = \
                    ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                        "ishareApiResponseData"][
                        "apiResponseData"][
                        0][
                        "responseMsg"]
                print(code)
                print(ishare_response)
                if code == '200' or ishare_response == 'Crediting Successful.':
                    return Response(data={"code": "0002"}, status=status.HTTP_200_OK)
                else:
                    pass
        else:
            pass
    else:
        pass
    print("moving on")
    data = saved_data
    history_web.collection(email).document(reference).set(data)

    if history_collection.document(reference).get().exists:
        print("first save")

    ishare_response, status_code = send_ishare_bundle(first_name=first_name, last_name=last_name, receiver=receiver,
                                                      buyer=phone,
                                                      bundle=data_volume,
                                                      email=email, reference=reference)
    json_response = ishare_response.json()
    print(f"hello:{json_response}")
    status_code = status_code
    print(status_code)
    try:
        batch_id = json_response["batch_id"]
    except KeyError:
        batch_id = "unknown"
    print(batch_id)

    doc_ref = history_collection.document(reference)
    if doc_ref.get().exists:
        doc_ref.update({'batch_id': batch_id, 'responseCode': status_code})
        history_web.collection(email).document(reference).update({'batch_id': batch_id, 'responseCode': status_code})
    else:
        print("didn't find any entry to update")
    print("firebase saved")
    # return status_code, batch_id if batch_id else "No batchId", email, first_name
    return Response(
        data=json_response,
        status=status.HTTP_200_OK)


def mtn_flexi_transaction(receiver, date, time, date_and_time, phone, amount, data_volume, details: dict, ref,
                          channel, txn_status):
    data = {
        'batch_id': "unknown",
        'buyer': phone,
        'color_code': "Green",
        'amount': amount,
        'data_break_down': str(data_volume),
        'data_volume': data_volume,
        'date': date,
        'date_and_time': date_and_time,
        'done': "unknown",
        'email': details["email"],
        'image': details["user_id"],
        'ishareBalance': 0,
        'name': f"{details['first_name']} {details['last_name']}",
        'number': receiver,
        'paid_at': date_and_time,
        'reference': ref,
        'responseCode': 200,
        'status': txn_status,
        'time': time,
        'tranxId': str(tranx_id_gen()),
        'type': "MTN Master Bundle",
        'uid': details["user_id"]
    }

    history_collection.document(date_and_time).set(data)
    history_web.collection(details['email']).document(date_and_time).set(data)
    user = history_collection.document(date_and_time)
    new_mtn_txn = models.MTNTransaction.objects.create(
        user_id=details["user_id"],
        amount=amount,
        bundle_volume=data_volume,
        number=receiver,
        firebase_date=date_and_time
    )
    new_mtn_txn.save()
    doc = user.get()
    print(doc.to_dict())
    tranx_id = doc.to_dict()['tranxId']
    second_data = {
        'amount': amount,
        'batch_id': "unknown",
        'channel': channel,
        'color_code': "Green",
        'created_at': date_and_time,
        'data_volume': data_volume,
        'date': date,
        'email': details["email"],
        'date_and_time': date_and_time,
        'image': details["user_id"],
        'ip_address': "",
        'ishareBalance': 0,
        'name': f"{details['first_name']} {details['last_name']}",
        'number': receiver,
        'buyer': phone,
        'paid_at': date_and_time,
        'payment_status': "success",
        'reference': ref,
        'status': txn_status,
        'time': time,
        'tranxId': tranx_id,
        'type': "MTN Master Bundle"
    }
    mtn_other.document(date_and_time).set(second_data)
    user22 = mtn_other.document(date_and_time)
    pu = user22.get()
    print(pu.to_dict())
    print("pu")
    mail_doc_ref = mail_collection.document()
    file_path = 'wallet_api_app/mtn_maill.txt'  # Replace with your file path

    name = details['first_name']
    volume = data_volume
    date = date_and_time
    reference_t = ref
    receiver_t = receiver

    with open(file_path, 'r') as file:
        html_content = file.read()

    placeholders = {
        '{name}': name,
        '{volume}': volume,
        '{date}': date,
        '{reference}': reference_t,
        '{receiver}': receiver_t
    }

    for placeholder, value in placeholders.items():
        html_content = html_content.replace(placeholder, str(value))

    mail_doc_ref.set({
        'to': details['email'],
        'message': {
            'subject': 'MTN Data',
            'html': html_content,
            'messageId': 'Bestpay'
        }
    })
    print("got to redirect")
    return Response(data={'code': '0000', 'message': "Transaction Saved"}, status=status.HTTP_200_OK)


def hubtel_mtn_flexi_transaction(saved_data, reference, email, data_volume, date_and_time, receiver, first_name):
    history_collection.document(reference).set(saved_data)
    history_web.collection(email).document(reference).set(saved_data)
    user = history_collection.document(reference)
    doc = user.get()
    print(doc.to_dict())
    tranx_id = doc.to_dict()['tranxId']
    second_data = saved_data
    mtn_other.document(reference).set(second_data)
    user22 = mtn_other.document(reference)
    pu = user22.get()
    print(pu.to_dict())
    print("pu")
    mail_doc_ref = mail_collection.document()
    file_path = 'wallet_api_app/mtn_maill.txt'  # Replace with your file path

    name = first_name
    volume = data_volume
    date = date_and_time
    reference_t = reference
    receiver_t = receiver

    with open(file_path, 'r') as file:
        html_content = file.read()

    placeholders = {
        '{name}': name,
        '{volume}': volume,
        '{date}': date,
        '{reference}': reference_t,
        '{receiver}': receiver_t
    }

    for placeholder, value in placeholders.items():
        html_content = html_content.replace(placeholder, str(value))

    mail_doc_ref.set({
        'to': email,
        'message': {
            'subject': 'MTN Data',
            'html': html_content,
            'messageId': 'Bestpay'
        }
    })
    print("got to redirect")
    return Response(data={'code': '0000', 'message': "Transaction Saved"}, status=status.HTTP_200_OK)


def big_time_transaction(receiver, date, time, date_and_time, phone, amount, data_volume, details: dict, ref,
                         channel, txn_status):
    data = {
        'batch_id': "unknown",
        'buyer': phone,
        'color_code': "Green",
        'amount': amount,
        'data_break_down': str(data_volume),
        'data_volume': data_volume,
        'date': date,
        'date_and_time': date_and_time,
        'done': "unknown",
        'email': details['email'],
        'image': details['user_id'],
        'ishareBalance': 0,
        'name': f"{details['first_name']} {details['last_name']}",
        'number': receiver,
        'paid_at': str(date_and_time),
        'reference': ref,
        'responseCode': 200,
        'status': txn_status,
        'time': time,
        'tranxId': str(tranx_id_gen()),
        'type': "AT Big Time",
        'uid': details['user_id']
    }
    history_collection.document(date_and_time).set(data)
    history_web.collection(details['email']).document(date_and_time).set(data)
    big_time.document(date_and_time).set(data)
    user = history_collection.document(date_and_time)
    doc = user.get()
    print(doc.to_dict())
    tranx_id = doc.to_dict()['tranxId']
    mail_doc_ref = mail_collection.document()
    file_path = 'wallet_api_app/mtn_mail.txt'  # Replace with your file path

    name = details['first_name']
    volume = data_volume
    date = date_and_time
    reference_t = ref
    receiver_t = receiver

    with open(file_path, 'r') as file:
        html_content = file.read()

    placeholders = {
        '{name}': name,
        '{volume}': volume,
        '{date}': date,
        '{reference}': reference_t,
        '{receiver}': receiver_t
    }

    for placeholder, value in placeholders.items():
        html_content = html_content.replace(placeholder, str(value))

    mail_doc_ref.set({
        'to': details['email'],
        'message': {
            'subject': 'Big Time Data',
            'html': html_content,
            'messageId': 'Bestpay'
        }
    })
    return Response(data={'code': '0000', 'message': "Transaction Saved"}, status=status.HTTP_200_OK)


def hubtel_big_time_transaction(saved_data, reference, email, data_volume, date_and_time, receiver, first_name):
    data = saved_data
    history_collection.document(reference).set(data)
    history_web.collection(email).document(reference).set(saved_data)
    big_time.document(reference).set(saved_data)
    user = history_collection.document(reference)
    doc = user.get()
    print(doc.to_dict())
    tranx_id = doc.to_dict()['tranxId']
    mail_doc_ref = mail_collection.document()
    file_path = 'wallet_api_app/mtn_mail.txt'  # Replace with your file path

    name = first_name
    volume = data_volume
    date = date_and_time
    reference_t = reference
    receiver_t = receiver

    with open(file_path, 'r') as file:
        html_content = file.read()

    placeholders = {
        '{name}': name,
        '{volume}': volume,
        '{date}': date,
        '{reference}': reference_t,
        '{receiver}': receiver_t
    }

    for placeholder, value in placeholders.items():
        html_content = html_content.replace(placeholder, str(value))

    mail_doc_ref.set({
        'to': email,
        'message': {
            'subject': 'Big Time Data',
            'html': html_content,
            'messageId': 'Bestpay'
        }
    })
    return Response(data={'code': '0000', 'message': "Transaction Saved"}, status=status.HTTP_200_OK)


def confirm(reference):
    import requests

    url = f"https://console.bestpaygh.com/api/flexi/v1/transaction_detail/{reference}/"

    payload = ""
    headers = {
        'api-key': config('api_key'),
        'api-secret': config('api_secret'),
    }

    response = requests.request("POST", url, headers=headers, data=payload)

    return response.json()


@csrf_exempt
def paystack_webhook(request):
    if request.method == "POST":
        paystack_secret_key = config("PAYSTACK_SECRET_KEY")
        # print(paystack_secret_key)
        payload = json.loads(request.body)

        paystack_signature = request.headers.get("X-Paystack-Signature")

        if not paystack_secret_key or not paystack_signature:
            return HttpResponse(status=400)

        computed_signature = hmac.new(
            paystack_secret_key.encode('utf-8'),
            request.body,
            hashlib.sha512
        ).hexdigest()

        if computed_signature == paystack_signature:
            print("yes")
            print(payload.get('data'))
            r_data = payload.get('data')
            print(r_data.get('metadata'))
            print(payload.get('event'))
            if payload.get('event') == 'charge.success':
                metadata = r_data.get('metadata')
                receiver = metadata.get('receiver')
                bundle_package = metadata.get('bundle_package')
                channel = metadata.get('channel')
                user_id = metadata.get('user_id')
                real_amount = metadata.get('real_amount')
                print(real_amount)
                paid_amount = float(r_data.get('amount')) / 100
                amount = real_amount
                email = r_data.get('email')
                reference = r_data.get('reference')
                date = metadata.get("date")
                time = metadata.get("time")
                date_and_time = metadata.get("date_and_time")
                txn_status = metadata.get("txn_status")

                user_details = get_user_details(user_id)
                if user_details is not None:
                    first_name = user_details['first name']
                    last_name = user_details['last name']
                    email = user_details['email']
                    phone = user_details['phone']
                    first_name = first_name
                else:
                    first_name = ""
                    email = ""

                if channel == "ishare":
                    send_response = webhook_send_and_save_to_history(user_id=user_id, date_and_time=date_and_time,
                                                                     date=date,
                                                                     time=time,
                                                                     amount=amount, receiver=receiver,
                                                                     reference=reference,
                                                                     paid_at=date_and_time,
                                                                     txn_type="AT Premium Bundle",
                                                                     color_code="Green", data_volume=bundle_package,
                                                                     ishare_balance=0, txn_status=txn_status)
                    data = send_response.data
                    json_response = data
                    print(json_response)
                    if json_response["code"] == "0002":
                        return HttpResponse(status=200)
                    elif json_response["code"] == "0001":
                        return HttpResponse(status=500)
                    else:
                        print(send_response.status_code)
                        try:
                            batch_id = json_response["batch_id"]
                        except KeyError:
                            return HttpResponse(status=500)

                        print(batch_id)

                        if json_response["code"] == '0000':
                            sms = f"Hey there\nYour account has been credited with {bundle_package}MB.\nConfirm your new balance using the AT Mobile App"
                            r_sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={receiver}&from=Bestpay&sms={sms}"
                            response = requests.request("GET", url=r_sms_url)
                            doc_ref = history_collection.document(date_and_time)
                            if doc_ref.get().exists:
                                doc_ref.update({'done': 'Successful'})
                            else:
                                print("no entry")
                            mail_doc_ref = mail_collection.document(f"{batch_id}-Mail")
                            file_path = 'wallet_api_app/mail.txt'  # Replace with your file path

                            name = first_name
                            volume = bundle_package
                            date = date_and_time
                            reference_t = reference
                            receiver_t = receiver

                            with open(file_path, 'r') as file:
                                html_content = file.read()

                            placeholders = {
                                '{name}': name,
                                '{volume}': volume,
                                '{date}': date,
                                '{reference}': reference_t,
                                '{receiver}': receiver_t
                            }

                            for placeholder, value in placeholders.items():
                                html_content = html_content.replace(placeholder, str(value))

                            mail_doc_ref.set({
                                'to': email,
                                'message': {
                                    'subject': 'AT Flexi Bundle',
                                    'html': html_content,
                                    'messageId': 'Bestpay'
                                }
                            })
                            print("donnee")
                            return HttpResponse(status=200)
                        else:
                            doc_ref = history_collection.document(date_and_time)
                            doc_ref.update({'done': 'Failed'})
                            return HttpResponse(status=500)
                elif channel == "mtn_flexi":
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }
                    mtn_response = mtn_flexi_transaction(receiver=receiver, date_and_time=date_and_time, date=date,
                                                         time=time, amount=amount, data_volume=bundle_package,
                                                         channel="MoMo", phone=phone, ref=reference, details=details,
                                                         txn_status=txn_status)
                    print("after mtn responses")
                    if mtn_response.status_code == 200 or mtn_response.data["code"] == "0000":
                        print("mtn donnnneeee")
                        print("yooo")
                        return HttpResponse(status=200)
                    else:
                        return HttpResponse(status=500)
                elif channel == "big-time":
                    if user_details is not None:
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }
                    big_time_response = big_time_transaction(receiver=receiver, date_and_time=date_and_time, date=date,
                                                             time=time, amount=amount, data_volume=bundle_package,
                                                             channel="MoMo", phone=phone, ref=reference,
                                                             details=details, txn_status=txn_status)
                    if big_time_response.status_code == 200 or big_time_response.data["code"] == "0000":
                        print("big time donnnneee")
                        return HttpResponse(status=200)
                    else:
                        return HttpResponse(status=500)
                elif channel == "top_up":
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        print(user_details)
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                        try:
                            previous_wallet = user_details['wallet']
                        except KeyError:
                            previous_wallet = 0
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                        previous_wallet = 0
                    all_data = {
                        'batch_id': "unknown",
                        'buyer': phone,
                        'color_code': "Green",
                        'amount': amount,
                        'data_break_down': amount,
                        'data_volume': bundle_package,
                        'date': date,
                        'date_and_time': date_and_time,
                        'done': "Success",
                        'email': email,
                        'image': user_id,
                        'ishareBalance': 0,
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'paid_at': date_and_time,
                        'reference': reference,
                        'responseCode': 200,
                        'status': txn_status,
                        'time': time,
                        'tranxId': str(tranx_id_gen()),
                        'type': "WALLETTOPUP",
                        'uid': user_id
                    }
                    history_web.collection(email).document(date_and_time).set(all_data)
                    print("f saved")
                    history_collection.document(date_and_time).set(all_data)
                    print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                    print("f saved")
                    print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")
                    to_be_added = float(amount)
                    print(f"amount to be added: {to_be_added}")
                    new_balance = previous_wallet + to_be_added
                    print(f" new balance: {new_balance}")
                    doc_ref = user_collection.document(user_id)
                    doc_ref.update(
                        {'wallet': new_balance, 'wallet_last_update': date_and_time,
                         'recent_wallet_reference': reference})
                    print(doc_ref.get().to_dict())
                    print("before all data")
                    all_data = {
                        'batch_id': "unknown",
                        'buyer': phone,
                        'color_code': "Green",
                        'amount': amount,
                        'data_break_down': amount,
                        'data_volume': bundle_package,
                        'date': date,
                        'date_and_time': date_and_time,
                        'done': "Success",
                        'email': email,
                        'image': user_id,
                        'ishareBalance': 0,
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'paid_at': date_and_time,
                        'reference': reference,
                        'responseCode': 200,
                        'status': txn_status,
                        'time': time,
                        'tranxId': str(tranx_id_gen()),
                        'type': "WALLETTOPUP",
                        'uid': user_id
                    }
                    history_web.collection(email).document(date_and_time).set(all_data)
                    print("saved")
                    history_collection.document(date_and_time).set(all_data)
                    print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                    print("saved")
                    print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")

                    name = f"{first_name} {last_name}"
                    amount = to_be_added
                    file_path = 'wallet_api_app/wallet_mail.txt'
                    mail_doc_ref = mail_collection.document()

                    with open(file_path, 'r') as file:
                        html_content = file.read()

                    placeholders = {
                        '{name}': name,
                        '{amount}': amount
                    }

                    for placeholder, value in placeholders.items():
                        html_content = html_content.replace(placeholder, str(value))

                    mail_doc_ref.set({
                        'to': email,
                        'message': {
                            'subject': 'Wallet Topup',
                            'html': html_content,
                            'messageId': 'Bestpay'
                        }
                    })

                    sms_message = f"GHS {to_be_added} was deposited in your wallet. Available balance is now GHS {new_balance}"
                    sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to=0{user_details['phone']}&from=Bestpay&sms={sms_message}"
                    response = requests.request("GET", url=sms_url)
                    print(response.status_code)
                    return HttpResponse(status=200)
                else:
                    return HttpResponse(status=200)
            else:
                return HttpResponse(status=200)
        else:
            return HttpResponse(status=401)
    else:
        return HttpResponse(status=405)


@csrf_exempt
def hubtel_webhook(request):
    if request.method == 'POST':
        try:
            payload = request.body.decode('utf-8')
            print("Hubtel payment Info: ", payload)
            json_payload = json.loads(payload)
            print(json_payload)

            data = json_payload.get('Data')
            print(data)
            reference = data.get('ClientReference')
            print(reference)
            txn_status = data.get('Status')
            amount = data.get('Amount')
            print(txn_status, amount)

            # all_data = {
            #     'batch_id': "unknown",
            #     'buyer': phone,
            #     'color_code': "Green",
            #     'amount': amount,
            #     'data_break_down': amount,
            #     'data_volume': bundle_package,
            #     'date': date,
            #     'date_and_time': date_and_time,
            #     'done': "Success",
            #     'email': email,
            #     'image': user_id,
            #     'ishareBalance': 0,
            #     'name': f"{first_name} {last_name}",
            #     'number': receiver,
            #     'paid_at': date_and_time,
            #     'reference': reference,
            #     'responseCode': 200,
            #     'status': txn_status,
            #     'time': time,
            #     'tranxId': str(tranx_id_gen()),
            #     'type': "WALLETTOPUP",
            #     'uid': user_id
            # }

            if txn_status == 'Success':
                print("success")
                collection_saved = history_collection.document(reference).get().to_dict()
                receiver = collection_saved['number']
                bundle_volume = collection_saved['data_volume']
                name = collection_saved['name']
                email = collection_saved['email']
                phone_number = collection_saved['buyer']
                date_and_time = collection_saved['date_and_time']
                txn_type = collection_saved['type']
                user_id = collection_saved['uid']
                print(receiver, bundle_volume, name, email, phone_number)

                doc_ref = history_collection.document(reference)

                if txn_type == "AT Premium Bundle":
                    doc_ref.update({'ishareBalance': "Paid", 'status': "Delivered", "tranxId": str(tranx_id_gen())})
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    collection_saved = history_collection.document(reference).get().to_dict()
                    send_response = hubtel_webhook_send_and_save_to_history(collection_saved, user_id, reference,
                                                                            receiver, bundle_volume)
                    # saved_data, user_id, reference, receiver, data_volume
                    data = send_response.data
                    json_response = data
                    print(json_response)
                    if json_response["code"] == "0002":
                        return HttpResponse(status=200)
                    elif json_response["code"] == "0001":
                        return HttpResponse(status=500)
                    else:
                        print(send_response.status_code)
                        try:
                            batch_id = json_response["batch_id"]
                        except KeyError:
                            return HttpResponse(status=500)

                        print(batch_id)

                        if json_response["code"] == '0000':
                            sms = f"Hey there\nYour account has been credited with {bundle_volume}MB.\nConfirm your new balance using the AT Mobile App"
                            r_sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={receiver}&from=Bestpay&sms={sms}"
                            response = requests.request("GET", url=r_sms_url)
                            doc_ref = history_collection.document(date_and_time)
                            if doc_ref.get().exists:
                                doc_ref.update({'done': 'Successful'})
                            else:
                                print("no entry")
                            mail_doc_ref = mail_collection.document(f"{batch_id}-Mail")
                            file_path = 'wallet_api_app/mail.txt'  # Replace with your file path

                            name = first_name
                            volume = bundle_volume
                            date = date_and_time
                            reference_t = reference
                            receiver_t = receiver

                            with open(file_path, 'r') as file:
                                html_content = file.read()

                            placeholders = {
                                '{name}': name,
                                '{volume}': volume,
                                '{date}': date,
                                '{reference}': reference_t,
                                '{receiver}': receiver_t
                            }

                            for placeholder, value in placeholders.items():
                                html_content = html_content.replace(placeholder, str(value))

                            mail_doc_ref.set({
                                'to': email,
                                'message': {
                                    'subject': 'AT Flexi Bundle',
                                    'html': html_content,
                                    'messageId': 'Bestpay'
                                }
                            })
                            print("donnee")
                            return JsonResponse({'message': "Success"}, status=200)
                        else:
                            doc_ref = history_collection.document(date_and_time)
                            doc_ref.update({'done': 'Failed'})
                            return JsonResponse({'message': "Success"}, status=200)
                elif txn_type == "MTN Master Data":
                    doc_ref.update({'ishareBalance': "Paid", 'status': "Undelivered", "tranxId": str(tranx_id_gen())})
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }
                    collection_saved = history_collection.document(reference).get().to_dict()
                    mtn_response = hubtel_mtn_flexi_transaction(collection_saved, reference, email, bundle_volume,
                                                                date_and_time, receiver, first_name)
                    print(mtn_response)
                    new_mtn_txn = models.MTNTransaction.objects.create(
                        user_id=user_id,
                        amount=amount,
                        bundle_volume=bundle_volume,
                        number=receiver,
                        firebase_date=date_and_time
                    )
                    new_mtn_txn.save()
                    # saved_data, reference, email, data_volume, date_and_time, receiver, first_name
                    print("after mtn responses")
                    if mtn_response.status_code == 200 or mtn_response.data["code"] == "0000":
                        print("mtn donnnneeee")
                        print("yooo")
                        return HttpResponse(status=200)
                    else:
                        return JsonResponse({'message': "Success"}, status=200)
                elif txn_type == "AT Big Time":
                    doc_ref.update({'ishareBalance': "Paid", 'status': "Undelivered", "tranxId": str(tranx_id_gen())})
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }
                    collection_saved = history_collection.document(reference).get().to_dict()
                    big_time_response = hubtel_big_time_transaction(collection_saved, reference, email, bundle_volume,
                                                                    date_and_time, receiver, first_name)
                    # saved_data, reference, email, data_volume, date_and_time, receiver, first_name
                    if big_time_response.status_code == 200 or big_time_response.data["code"] == "0000":
                        print("big time donnnneee")
                        return JsonResponse({'message': "Success"}, status=200)
                    else:
                        return HttpResponse(status=500)
                elif txn_type == "Bestpay E - Wallet":
                    doc_ref.update({'ishareBalance': "Paid", 'status': "Credited", "tranxId": str(tranx_id_gen())})
                    user_details = get_user_details(user_id)
                    collection_saved = history_collection.document(reference).get().to_dict()
                    if user_details is not None:
                        print(user_details)
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                        try:
                            previous_wallet = user_details['wallet']
                        except KeyError:
                            previous_wallet = 0
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                        previous_wallet = 0
                    all_data = collection_saved
                    history_web.collection(email).document(date_and_time).set(all_data)
                    print("f saved")
                    history_collection.document(date_and_time).set(all_data)
                    print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                    print("f saved")
                    print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")
                    to_be_added = float(amount)
                    print(f"amount to be added: {to_be_added}")
                    new_balance = previous_wallet + to_be_added
                    print(f" new balance: {new_balance}")
                    doc_ref = user_collection.document(user_id)
                    doc_ref.update(
                        {'wallet': new_balance, 'wallet_last_update': date_and_time,
                         'recent_wallet_reference': reference})
                    print(doc_ref.get().to_dict())
                    print("before all data")
                    all_data = collection_saved
                    history_web.collection(email).document(date_and_time).set(all_data)
                    print("saved")
                    print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                    print("saved")
                    print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")

                    name = f"{first_name} {last_name}"
                    amount = to_be_added
                    file_path = 'wallet_api_app/wallet_mail.txt'
                    mail_doc_ref = mail_collection.document()

                    with open(file_path, 'r') as file:
                        html_content = file.read()

                    placeholders = {
                        '{name}': name,
                        '{amount}': amount
                    }

                    for placeholder, value in placeholders.items():
                        html_content = html_content.replace(placeholder, str(value))

                    mail_doc_ref.set({
                        'to': email,
                        'message': {
                            'subject': 'Wallet Topup',
                            'html': html_content,
                            'messageId': 'Bestpay'
                        }
                    })

                    sms_message = f"GHS {to_be_added} was deposited in your wallet. Available balance is now GHS {new_balance}"
                    sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to=0{user_details['phone']}&from=Bestpay&sms={sms_message}"
                    response = requests.request("GET", url=sms_url)
                    print(response.status_code)
                    return JsonResponse({'message': "Success"}, status=200)
                else:
                    print("no type found")
                    return JsonResponse({'message': "No Type Found"}, status=500)
            else:
                doc_ref = history_collection.document(reference)
                doc_ref.update({'status': "Failed"})
                return JsonResponse({'message': 'Transaction Failed'}, status=200)
        except Exception as e:
            print("Error Processing hubtel webhook:", str(e))
            return JsonResponse({'status': 'error'}, status=500)
    else:
        print("not post")
        return JsonResponse({'message': 'Not Found'}, status=404)



from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import pandas as pd
from .models import MTNTransaction  # Adjust the import based on your model's location

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


from openpyxl import load_workbook

@csrf_exempt
def export_unknown_transactions(request):
    existing_excel_path = 'wallet_api_app/ALL PACKAGES LATEST.xlsx'  # Update with your file path

    # Load the existing Excel file using openpyxl.Workbook
    book = load_workbook(existing_excel_path)

    # Get the active sheet
    sheet_name = 'Sheet1'
    sheet = book[sheet_name] if sheet_name in book.sheetnames else book.active

    # Clear existing data from the sheet (excluding headers)
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    # Query your Django model for the first 200 records with batch_id 'Unknown' and ordered by status and date
    queryset = MTNTransaction.objects.filter(batch_id='Unknown', status="Undelivered")[:60]

    # Process transactions with batch_id 'Unknown'
    counter = 0

    for record in queryset:
        print(counter)

        # Extract required fields from your Django model
        bundle_volume_mb = record.bundle_volume  # Assuming a default of 0 if datavolume is missing
        number = str(record.number)  # Convert to string to keep leading zeros

        # Convert datavolume from MB to GB
        bundle_volume_gb = round(float(bundle_volume_mb) / 1000)

        # Find the row index where you want to populate the data (adjust as needed)
        target_row = 2 + counter  # Assuming the data starts from row 2

        # Populate the specific cells with the new data
        sheet.cell(row=target_row, column=1, value=number)  # Keep leading zeros
        sheet.cell(row=target_row, column=2, value=float(bundle_volume_gb))  # Convert to float

        # Update 'batch_id' to 'processing' in your Django model
        record.batch_id = 'accepted'
        record.status = 'Processing'
        record.save()

        counter += 1
        txn = mtn_other.document(record.firebase_date)
        txn.update({'batch_id': 'accepted', 'status': 'Processing'})

    print(f"Total transactions to export: {counter}")

    # Save changes to the existing Excel file
    book.save(existing_excel_path)

    # You can continue with the response as needed
    excel_buffer = BytesIO()

    # Save the workbook to the buffer
    book.save(excel_buffer)

    # Create a response with the Excel file
    response = HttpResponse(excel_buffer.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={datetime.datetime.now()}.xlsx'

    return response

